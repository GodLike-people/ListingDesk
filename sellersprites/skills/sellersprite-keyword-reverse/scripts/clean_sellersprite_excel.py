from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REQUIRED_HEADERS = [
    "关键词",
    "预估周曝光量",
    "关键词类型",
    "流量词类型",
    "自然流量占比",
    "广告流量占比",
    "自然排名",
    "自然排名页码",
    "广告排名",
    "广告排名页码",
    "ABA周排名",
    "月搜索量",
    "SPR",
    "建议竞价范围",
]


KNOWN_COLUMN_FALLBACK = {
    "关键词": 1,
    "预估周曝光量": 4,
    "关键词类型": 5,
    "流量词类型": 7,
    "自然流量占比": 8,
    "广告流量占比": 9,
    "自然排名": 10,
    "自然排名页码": 11,
    "广告排名": 13,
    "广告排名页码": 14,
    "ABA周排名": 16,
    "月搜索量": 17,
    "SPR": 18,
    "建议竞价范围": 30,
}


def normalize(value: object) -> str:
    return re.sub(r"[\s_\-]+", "", str(value or "")).lower()


def resolve_columns(sheet) -> list[int]:
    actual = {
        normalize(sheet.cell(1, column).value): column
        for column in range(1, sheet.max_column + 1)
    }
    resolved: list[int] = []
    missing: list[str] = []
    for header in REQUIRED_HEADERS:
        column = actual.get(normalize(header))
        if column is None:
            fallback = KNOWN_COLUMN_FALLBACK[header]
            if fallback <= sheet.max_column:
                column = fallback
            else:
                missing.append(header)
                continue
        resolved.append(column)
    if missing:
        raise ValueError("Missing required columns: " + ", ".join(missing))
    return resolved


def clean(source: Path, output_directory: Path) -> tuple[Path, int, int]:
    source_book = load_workbook(source, read_only=True, data_only=True)
    source_sheet = source_book[source_book.sheetnames[0]]
    columns = resolve_columns(source_sheet)

    output_book = Workbook()
    output_sheet = output_book.active
    output_sheet.title = "关键词清洗"
    output_sheet.append(REQUIRED_HEADERS)

    rows_written = 0
    for source_row in source_sheet.iter_rows(min_row=2, values_only=True):
        values = [source_row[column - 1] for column in columns]
        if not any(value not in (None, "") for value in values):
            continue
        output_sheet.append(values)
        rows_written += 1

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in output_sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column in (5, 6):
        for (cell,) in output_sheet.iter_rows(
            min_row=2, min_col=column, max_col=column
        ):
            cell.number_format = "0.00%"

    widths = [32, 16, 18, 22, 16, 16, 12, 20, 12, 20, 14, 14, 10, 18]
    for index, width in enumerate(widths, start=1):
        output_sheet.column_dimensions[get_column_letter(index)].width = width

    output_sheet.freeze_panes = "A2"
    output_sheet.auto_filter.ref = output_sheet.dimensions
    output_sheet.row_dimensions[1].height = 24

    output_directory.mkdir(parents=True, exist_ok=True)
    destination = output_directory / f"{source.stem}-cleaned.xlsx"
    output_book.save(destination)
    source_book.close()
    return destination, rows_written, len(REQUIRED_HEADERS)


if __name__ == "__main__":
    if len(sys.argv) != 3:
        raise SystemExit(
            "usage: clean_sellersprite_excel.py SOURCE.xlsx OUTPUT_DIRECTORY"
        )
    result, rows, columns = clean(Path(sys.argv[1]), Path(sys.argv[2]))
    print(f"output={result}")
    print(f"rows={rows} columns={columns}")
